import datetime
import requests
from excel_funcs import AddRates, SaveExcel
from openpyxl.workbook import Workbook
from helpers import getAllRates as getAllRates

from config import DAYS

from rent_motors_parser import getOneColumnRates as rent_motors_func
from rex_rent_parser import getOneColumnRates as rex_rent_func
from a_arenda_parser import getOneColumnRates as a_arenda_func
from almak_parser import getOneColumnRates_almak as almak_func
from storlet_car_parser import getOneColumnRates as storlet_func

import asyncio
import aiohttp

from typing import List
from objects import CFR

async def main():
    async with aiohttp.ClientSession() as s:
        
        strat_time = datetime.datetime.today()

        for d in DAYS:
            print(f'>>> Rental start date after {d} days from taday..')
            pick_up = datetime.datetime.today() + datetime.timedelta(d)

            # SET UP EXCELL INFO SHEET
            wb = Workbook()
            ws = wb.active
            ws.title = "INFO"
            ws.cell(1, 1).value = "PARSER 1.0 by fckklubb"
            ws.cell(3, 1).value = f"Parsing date: {datetime.datetime.today().strftime('%d-%m-%Y')}"
            ws.cell(4, 1).value = f"Start date of a rental: {pick_up.strftime('%d-%m-%Y')}"
            ws.cell(6, 1).value = "Next companies are going to be parsed.."
            
            res_list: List[CFR]

            # Rent Motors
            print("1. Rent Motors..")
            ws.cell(8, 1).value = "RENT MOTORS"
            res_list = await getAllRates(pick_up, s, rent_motors_func, location=10)
            wb = AddRates("RENT-MOTORS", wb, res_list)
            if wb == None:
                print("Panic.. Rent Motors failed..")

            # Rex Rent
            print("2. Rex Rent..")
            ws.cell(9, 1).value = "REX RENT"
            res_list = await getAllRates(pick_up, s, rex_rent_func)
            wb = AddRates("REX-RENT", wb, res_list)
            if wb == None:
                print("Panic.. Rex Rent failed..")

            # A Arenda
            print("3. A Arenda..")
            ws.cell(10, 1).value = "A ARENDA"
            res_list = await getAllRates(pick_up, s, a_arenda_func)
            wb = AddRates("A-ARENDA", wb, res_list)
            if wb == None:
                print("Panic.. A Arenda failed..")

            # Almak
            print("4. Almak..")
            ws.cell(11, 1).value = "ALMAK"
            res_list = await getAllRates(pick_up, s, almak_func, loop=False, xtra=False)
            wb = AddRates("ALMAK", wb, res_list)
            if wb == None:
                print("Panic.. Almak failed..")

            # Storlet
            print("5. Storlet..")
            ws.cell(12, 1).value = "STORLET"
            res_list = await getAllRates(pick_up, s, storlet_func)
            wb = AddRates("STORLET", wb, res_list)


            print("All done..")

            SaveExcel(wb, add_time=True)
            print(f"Start @ {strat_time}, finish @ {datetime.datetime.today()}")

if __name__ == '__main__':
    asyncio.run(main())