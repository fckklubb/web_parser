from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from typing import List
from objects import CFR
import datetime

def SaveExcel(wb: Workbook, add_time: bool):
    if not add_time:
        wb.save(f"outputs/Competitors-rates-{datetime.datetime.today().strftime('%d-%m-%Y')}.xlsx")
    else:
        wb.save(f"outputs/Competitors-rates-{datetime.datetime.today().strftime('%d-%m-%Y-%H_%M')}.xlsx")

def AddRates(company: str, wb: Workbook, data: List[CFR]) -> Workbook:
    if data == None:
        return None
    ws = wb.create_sheet(f"{company}-{datetime.datetime.today().strftime('%d-%m-%y')}")
    firstColumn = ['SIPP', 'CAR NAME', 'MILLAGE', 'DEPO', 'R1', 'R3', 'R7', 'R14', 'R21', 'R30']
    if not data[0].different_millages:
        for i in range(0, len(firstColumn)):
            ws.cell(i+1, 1).value = firstColumn[i]
        for i in range(0, len(data)):
            arr = [data[i].sipp, data[i].name, data[i].millage, data[i].depo, data[i].rate1, data[i].rate3, data[i].rate7, data[i].rate14, data[i].rate21, data[i].rate30]
            for j in range(0, len(arr)):
                ws.cell(j+1, i+2).value = arr[j]
    else:
        for i in range(0, len(firstColumn)):
            ws.cell(i+1, 1).value = firstColumn[i]
            ws.cell(i+12, 1).value = firstColumn[i]
            ws.cell(i+23, 1).value = firstColumn[i]
            c = len(data)//3
        for i in range(0, len(data)):
            arr = [data[i].sipp, data[i].name, data[i].millage, data[i].depo, data[i].rate1, data[i].rate3, data[i].rate7, data[i].rate14, data[i].rate21, data[i].rate30]
            for j in range(0, len(arr)):
                if data[i].millage == 200:
                    ws.cell(j+1, i+2).value = arr[j]
                if data[i].millage == 300:
                    ws.cell(j+12, i+2-c).value = arr[j]
                if data[i].millage == 500:
                    ws.cell(j+23, i+2-c*2).value = arr[j]
    return wb