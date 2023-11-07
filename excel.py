import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet_test'
ws.cell(1,1).value = "12"
ws.cell(1,2).value = "13"
wb.save("test.xlsx")